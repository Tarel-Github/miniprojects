# 대량 메일 전송
# pip install openpyxl
from openpyxl import load_workbook
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart

wbook = load_workbook('./studyPython/SpamMailList.xlsx', data_only=True)
wsheet = wbook.active # sheet1 선택

for i in range(1, wsheet.max_row+1):
    recv_mail = wsheet.cell(i, 1).value # 받는 사람이름은 엑셀에서 가져옴
    print(recv_mail)
    try:
        # 실제 메일전송 로직
        send_mail = 'sms_stmax@naver.com'
        send_pass = '135xldpa&*%%'
        smtp_name = 'smtp.naver.com'
        smtp_port = 587
        msg =MIMEMultipart()
        msg['Subject'] = '카카오 가고싶다.'
        msg['From'] = send_mail
        msg['To'] = recv_mail
        msg.attach(MIMEText('그/아/아/앗!!'))

        mail = smtplib.SMTP(smtp_name, smtp_port) # 객체생성
        mail.starttls() # 보안
        mail.login(send_mail, send_pass)
        mail.sendmail(send_mail, recv_mail, msg.as_string())
        mail.quit()
        print(f'전송성공 : {recv_mail}')
    except Exception as e:
        print(f'수신메일 - {recv_mail}')
        print(f'전송에러 : {e}')
        